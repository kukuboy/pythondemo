# -*- codeing = utf-8 -*-
# @Time : 2021/4/21 15:09
# @Author : 水印红枫
# @Software: PyCharm


from openpyxl import load_workbook  # 进行excel操作
import urllib.request  # 指定url获取网络数据
from bs4 import BeautifulSoup  # 网页解析获取数据


def main():
    base_url = "https://fundf10.eastmoney.com/F10DataApi.aspx?type=lsjz&sdate=2001-12-18&per=20"
    code_dict = dict()
    # 读取代码
    print("----------开始读取数据----------")
    try:
        wordbook = load_workbook("天天基金网  数据导出.xlsx")
        worksheet = wordbook.active
        max_column = worksheet.max_column
        col = 2
        while col <= max_column:
            code = worksheet.cell(row=2, column=col).value
            code_dict[code] = []
            col += 1
        # print(code_dict)
    except Exception as error:
        print("读取出错，原因为：", error)
    # 爬取数据
    print("----------开始爬取数据----------")
    for code in code_dict:
        def page_add(t_url, t_page):
            print("正在获取%s的第%d页的数据" % (code, t_page))
            try:
                data = getData(t_url + "&page=" + str(t_page))
                bs = BeautifulSoup(data, "html.parser")
                bs_list = bs.find("table").find_all("tr")
                for tr in bs_list:
                    tr_list = []
                    for td in tr.find_all("td"):
                        tr_list.append(td.get_text())
                    if len(tr_list):
                        code_dict[code].append(tr_list)
                # print(len(bs_list), code_dict)
                if len(bs_list) > 20:
                    page_add(t_url, t_page + 1)
            except Exception as error_add:
                print("处理出错，原因为：", error_add, data)

        url = base_url + "&code=" + code
        page_add(url, 1)
    # 写入数据
    # print(code_dict)
    print("----------开始写入数据----------")
    # 找到最长的那个基金
    max_l = 0
    max_code = ""
    for code in code_dict:
        if len(code_dict[code]) > max_l:
            max_l = len(code_dict[code])
            max_code = code
    # 将最长基金的时间倒序写入表格
    print(max_l)
    row_index = 3
    while max_l > 0:
        max_l -= 1
        worksheet.cell(row=row_index, column=1).value = code_dict[max_code][max_l][0]
        row_index += 1

    # 按值查找行
    def getRowByCol(val):
        max_row = worksheet.max_row
        while max_row > 0:
            if worksheet.cell(row=max_row, column=1).value == val:
                break
            max_row -= 1
        return max_row

    # 依次写入数据
    column_index = 2
    for code in code_dict:
        print("正在写入%s的数据" % code)
        for code_data in code_dict[code]:
            try:
                row_index = getRowByCol(code_data[0])
                if row_index > 0:
                    worksheet.cell(row=row_index, column=column_index).value = code_data[3]
            except Exception as error_list:
                print("处理出错，原因为：", error_list, code_data)
        column_index += 1
    # 保存数据
    wordbook.save("天天基金网  数据导出-备份.xlsx")


def getData(url):
    response_data = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36"
    }
    try:
        req = urllib.request.Request(url=url, headers=headers)
        response = urllib.request.urlopen(req, timeout=10)
        response_data = response.read().decode("utf-8")
    except Exception as e:
        if hasattr(e, "reason"):
            print(e.reason)
    return response_data


if __name__ == "__main__":
    main()
    print("程序结束")
