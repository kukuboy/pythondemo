# -*- codeing = utf-8 -*-
# @Time : 2021/4/7 13:53
# @Author : 水印红枫
# @Software: PyCharm

import openpyxl

# 加载获取工作簿
workbook = openpyxl.load_workbook(r"D:\www\mine\python\demo\excel\excelFrom.xlsx", data_only=True)
# workbook = openpyxl.load_workbook("excelFrom.xlsx")

# 创建新的工作表
workbook.create_sheet("这是创建的")
workbook.create_sheet("这是删除的")

# 删除工作表
workbook.remove(workbook["这是删除的"])

# 复制工作表
workbook_copy = workbook.copy_worksheet(workbook["这是创建的"])

# 更改表名
workbook_copy.title = "这是复制的"

# 获取当前激活的工作表，为上次保存时激活的工作表
worksheet = workbook.active

# 获取第一个工作表
# worksheet = workbook[0]
# 获取置顶姓名的工作表
# worksheet = workbook["Sheet1"]

# 获取区域内的单元格
# 'A1:B2' 对角区域 '1:10'第一行到第十行 'A:C'第一列到第三列
worksheet_area = worksheet['A1:B2']  # 获取到的是整个元组-》每一行的元组->每个单元格
# worksheet_area_list = list(worksheet_area) # 一次将每一行的元组放入到列表中
print(worksheet, worksheet_area, sep="\n")

# 获取所有的工作表，以列表形式返回
worksheets = workbook.worksheets
print(worksheets)

# 返回工作表的名称
print(worksheets[0].title)

# 获取单元格，获取单元格内容,获取单元格的行，单元格的列
cell = worksheet["A1"]
cell1 = worksheet.cell(row=1, column=1)
# cell1 = worksheet.cell(1, 1) #这种方法行在前，列在后
print(cell.row, cell.column, cell, cell.value, cell1, cell1.value, sep="\n")

# 遍历工作表
rows = worksheet.iter_rows(min_row=3, max_row=5, min_col=2, max_col=3)

# 元组形式遍历
# rows = worksheet.rows
# columns = worksheet.columns
for row in rows:
    # print(row[0].value)
    for cell in row:
        # print(cell.value)
        pass
# 数组转字母,字母转数字

print(openpyxl.utils.get_column_letter(2))
print(openpyxl.utils.column_index_from_string('B'))

# 获取最大行，最大列
print(worksheet.max_row, worksheet.max_column)

# 单元格修改
worksheet['G1'] = "添加的"
worksheet.cell(10, 10, value="添加的")
# cell.value = "添加的"
# worksheet.append(['1', '2', '3']) # 添加到最后一行

# 在第几行，第几列插入，删除几个 ,idx从哪开始，amount总共几个
# worksheet.insert_rows(1,5)
# worksheet.insert_cols(1,5)
# worksheet.delete_rows(1, 5)
# worksheet.delete_cols(1, 5)

# 将某个区域的单个格向上下左右移动，负值即为向上和向左
# worksheet.move_range('A1:B2', rows=2, cols=1)

# 冻结单元格，可用来做表头和聊表名
# worksheet.freeze_panes = "C3"

# 合并与分开单元格
# worksheet.merge_cells('A1:B2')
# worksheet.unmerge_cells('A1:B2')

# 使用公式
# worksheet['E3'] = "=sum(A3:b3)"

# 合并行或者列
# worksheet.row_dimensions.group(1, 1, hidden=True)
# worksheet.column_dimensions.group((1, 1))

# 增加批注
comment = openpyxl.comments.Comment("这是批注内容", "这是批注的人")
worksheet['A1'].comment = comment
# 保存工作表
workbook.save("excelFrom.xlsx")
