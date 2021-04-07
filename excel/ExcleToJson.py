# -*- codeing = utf-8 -*-
# @Time : 2021/4/7 10:44
# @Author : 水印红枫
# @Software: PyCharm


from openpyxl import load_workbook  # 进行excel操作
import json  # 转换为json


def main():
    # name = r"D:\file\wechat\WeChat Files\wxid_k2bas56p2sbz21\FileStorage\File\2021-04\中奖名单公示链接"
    name = "中奖名单公示链接"
    end = "xlsx"
    val = dict()
    try:
        wordbook = load_workbook(name + "." + end)
        table = wordbook["Sheet1"]
        rows = table.iter_rows
        val["table1"] = []
        val["table2"] = []
        for r in rows(min_row=3, max_row=7):
            row = dict()
            row["count"] = r[1].value
            row["phone"] = r[2].value
            row["prize"] = r[3].value
            val["table1"].append(row)
        for r in rows(min_row=10):
            row = dict()
            row["count"] = r[1].value
            row["phone"] = r[2].value
            row["prize"] = r[3].value
            val["table2"].append(row)
    except Exception as error:
        print("读取出错，原因为：", error)
    # 保存到json
    try:
        print("---------------正在保存到json")
        f = open(name + ".json", "w", encoding="utf-8")
        try:
            f.write(json.dumps(val, ensure_ascii=False, indent=2))
        finally:
            f.close()
    except Exception as error:
        print("保存到文件出错，原因为：", error)


if __name__ == "__main__":
    main()
    print("程序结束")
